""" Этот scripts осуществляет через Bitrix24 API поиск кандидата в базе по id,
сохраняет данные в файл и прикрепляет ссылку на файл к кандидату.
Также возможность создания смарт процесса для Bitrix24, которая получает необохдимые данные при
загрузки из файла. 

Импортированные модули:
- openpyxl: Модуль для работы с файлами Excel (.xlsx). 
- requests: Модуль для для выполнения HTTP-запросов.
- datetime: Модуль предоставляет классы для работы с датами и временем.
- logging: Модуль для логирования. 
"""


import openpyxl
from openpyxl.styles import Font, Alignment
import requests
from datetime import datetime
import logging


# Настройка логгирования. Данные хранятся в файле bitrix24.log с уровнем логирования INFO. 
logging.basicConfig(
    filename = 'bitrix24.log',  
    level = logging.INFO,  
    format = '%(asctime)s - %(levelname)s - %(message)s', 
)


# Настройки констант для Bitrix24 API 
# (WEBHOOK - должен предоставлять доступ к определенным функциям)
BITRIX_WEBHOOK_URL = 'https://your_domain.bitrix24.ru/rest/1/your_webhook/'


def get_candidate_data(candidate_id: int) -> dict:
    """Получает данные кандидата из системы Bitrix24 по уникальному идентификатору.
    
    :param candidate_id: уникальный идентификатор кандидата.
    :return: словарь с информацией о кандидате, или сообщение об ошибке.
    """
    
    logging.info(f'start get_candidate_data - {candidate_id}')
    
    url = f"{BITRIX_WEBHOOK_URL}crm.lead.get"
    params = {"id": candidate_id}
    
    try: 
        response = requests.get(url, params=params)
        response.raise_for_status()  # проверка успешности ответа

        logging.info(f'Candidate details received')
        try:
            candidate_data = response.json()
            return candidate_data  # Возвращаем данные в случае успешного запроса
        except ValueError as json_error:
            logging.error("Error decoding JSON. Response text: %s", response.text)
            logging.exception("JSON decode error: %s", json_error)
            return {"error": "Error decoding JSON from response."}
        
    except requests.exceptions.HTTPError as http_error:
        logging.error("HTTP error: %s, Status code: %d", http_error, http_error.response.status_code)
        return {"error": f"HTTP error: {http_error}"}
    except requests.exceptions.ConnectionError:
        logging.error("Connection error occurred while attempting to reach Bitrix24.")
        return {"error": "Connection error occurred."}
    except requests.exceptions.Timeout:
        logging.error("The request timed out.")
        return {"error": "Request timed out."}
    except requests.exceptions.RequestException as req_error:
        logging.error("Request exception: %s", req_error)
        return {"error": f"Request failed: {req_error}"}
    except Exception as e:
        logging.error("An unexpected error occurred: %s", e)
        return {"error": "An unexpected error occurred."}

def save_candidate_to_excel(candidate_data: dict, file_name: str = 'candidates') -> str:
    """Сохраняет данные о кандидате в файл формата Excel (*.xlsx*).
    
    :param candidate_data: словарь с данными кандидата.
    :param file_name: имя файла, куда будут сохраняться данные. По умолчанию используется значение 'candidates'.
    :return: сообщение о наличии сохраненных данных или об ошибке.
    """
    try: 
        logging.info(f'start save_candidate_to_excel')
        
        # Проверяем, есть ли необходимые данные в candidate_data
        if 'result' not in candidate_data:
            logging.warning("No 'result' key found in candidate data.")
            return "No candidate data to save."
        
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Кандидаты'
        
        # Стили заголовков
        header_font = Font(bold=True)
        align_center = Alignment(horizontal='center')
        
        # Заголовки столбцов
        headers = ['ID', 'Имя', 'Фамилия', 'Телефон', 'Email', 'Дата создания']
        sheet.append(headers)
        for col, value in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = value
            cell.font = header_font
            cell.alignment = align_center

        # Заполнение данных
        row_num = 2
        result = candidate_data['result']
        sheet.cell(row=row_num, column=1).value = result.get('ID', 'N/A')
        sheet.cell(row=row_num, column=2).value = result.get('NAME', 'N/A')
        sheet.cell(row=row_num, column=3).value = result.get('LAST_NAME', 'N/A')
        
        phone_list = result.get('PHONE', [])
        email_list = result.get('EMAIL', [])
        
        sheet.cell(row=row_num, column=4).value = phone_list[0]['VALUE'] if phone_list else 'N/A'
        sheet.cell(row=row_num, column=5).value = email_list[0]['VALUE'] if email_list else 'N/A'
        
        sheet.cell(row=row_num, column=6).value = datetime.strptime(result.get('DATE_CREATE', ''), '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y') if 'DATE_CREATE' in result else 'N/A'
        
        # Сохранение файла
        file_name_save = f"{file_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        workbook.save(file_name_save)
        logging.info(f'Candidate data saved in {file_name_save}')
        return f"Данные кандидата сохранены в {file_name_save}"
    
    except FileNotFoundError as fnf_error:
        logging.error("File not found: %s", fnf_error)
        return "Error: File not found."
    except ValueError as ve:
        logging.error("Value error occurred: %s", ve)
        return "Error: Problem with data values."
    except Exception as e:
        logging.error("Error save: %s", e)
        return f"Failed to save candidate data: {str(e)}"


def upload_file_to_lead(file_name: str) -> int:
    """Создание поля в карточке кандидата Bitrix24 для прикрепления файла.

    :param file_name: имя файла, который необходимо прикрепить к карточке
    :return: id созданного поля или сообщение об ошибке.
    """
    
    logging.info('Start upload_file_to_lead')
    
    url = f'{BITRIX_WEBHOOK_URL}crm.lead.userfield.add'  # CRM для создания поля
    try:
        # Подготовка полезной нагрузки
        payload = {
            "fields": {
                "FIELD_NAME": "LINK_TO_CANDIDATS",
                "EDIT_FORM_LABEL": f"Ссылка на файл {file_name.split('_')[0]}",
                "LIST_COLUMN_LABEL": f"Ссылка на файл '{file_name.split('_')[0]}'",
                "USER_TYPE_ID": "file",
                "MULTIPLE": "N",
                "MANDATORY": "N",
                "SHOW_FILTER": "N",
                "SHOW_IN_LIST": "Y",
                "IS_SEARCHABLE": "N",
                "SORT": 100,
                "XML_ID": "LINK_TO_CANDIDATE_FILE"
            }
        }

        # Отправка запроса
        response = requests.post(url, json=payload)
        response.raise_for_status()  # Проверка успешности запроса

        result = response.json()

        if 'result' in result:
            logging.info(f"Поле успешно создано с ID: {result['result']}")
            return result['result']
        else:
            logging.error(f"Failed to create field: {result}")
            return None

    except requests.exceptions.HTTPError as http_error:
        logging.error(f"HTTP error occurred: {http_error} - Status code: {http_error.response.status_code}")
        return None
    except requests.exceptions.RequestException as req_error:
        logging.error(f"Request exception occurred: {req_error}")
        return None
    except Exception as e:
        logging.error(f"Ошибка при создании поля: {e}")
        return None


def save_link_to_file(field_id: int, file_path: str, candidate_id: int):
    """Прикрепление ссылки на файл с данными кандидата к карточке кандидата Bitrix24.

    :param field_id: id поля для прикрепления ссылки.
    :param file_path: путь к файлу.
    :param candidate_id: id кандидата, к которому прикрепить файл.
    """
    
    logging.info('Start save_link_to_file')
    
    url = f'{BITRIX_WEBHOOK_URL}crm.lead.update.json'  # CRM для обновления поля
    try: 
        payload = {
            "id": candidate_id,
            "fields": {
                field_id: {"value": file_path}
            }
        }

        # Отправка запроса
        response = requests.post(url, json=payload)
        response.raise_for_status()  # Проверка успешности запроса
        
        result = response.json()
        
        if 'result' in result:
            logging.info(f"Ссылка на файл успешно сохранена у кандидата - {candidate_id}")
        else:
            logging.error(f"Failed to save file link for candidate {candidate_id}: {result}")

    except requests.exceptions.HTTPError as http_error:
        logging.error(f"HTTP error while saving link: {http_error} - Status code: {http_error.response.status_code}")
    except requests.exceptions.RequestException as req_error:
        logging.error(f"Request exception occurred while saving link: {req_error}")
    except Exception as e:
        logging.error(f"Ошибка при сохранении ссылки на файл: {e}")

def read_from_excel(file_name: str) -> list:
    """ Функция предназначена чтения файла формата Excel (*.xlsx*)
    
    :param file_name: имя файла, который необходимо открыть
    :return: список с данными из файла.
    """
    
    logging.info('Start reading from Excel file: %s', file_name)
    try:
        # Открытие книги
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        
        data = []
        headers = next(ws.iter_rows(values_only=True), None)  # Пропустить первую строку с заголовками
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:  # Проверка на количество колонок
                logging.warning("Row has insufficient columns: %s", row)
                continue  # Пропускаем некорректные строки
            data.append({
                'TITLE': row[0], 
                'NAME': row[1],  
                'LAST_NAME': row[2],  
                'PHONE': row[3],  
                'EMAIL': row[4]
            })
        logging.info('File read successfully: %s', file_name)
        return data
    except FileNotFoundError:
        logging.error("File not found: %s", file_name)
        return []
    except openpyxl.utils.exceptions.InvalidFileException:
        logging.error("Invalid file format: %s", file_name)
        return []
    except Exception as e:
        logging.error("Error reading Excel file: %s", e)
        return []

def create_smart_process(data: list):
    """Функция предназначена создания смарт процесса в Bitrix24.
    
    :param data: данные для загрузки смарт процесса (list).
    """
    
    logging.info('Starting to create smart processes')
    
    if not data:
        logging.warning("No data provided to create smart processes.")
        return
    
    smart_process_url = f'{BITRIX_WEBHOOK_URL}crm.lead.add'  # CRM для создания процесса

    for item in data:
        if not ('TITLE' in item and 'LAST_NAME' in item):
            logging.warning("Item missing required fields: %s", item)
            continue
        
        lead_data = {
            'fields': {
                'TITLE': item['TITLE'],
                'NAME': item['NAME'] if item['NAME'] else 'Empty name',
                'LAST_NAME': item['LAST_NAME'],
                'PHONE': [{'VALUE': item['PHONE'], 'VALUE_TYPE': 'HOME'}] if item['PHONE'] else [],
                'EMAIL': [{'VALUE': item['EMAIL'], 'VALUE_TYPE': 'HOME'}] if item['EMAIL'] else []
            }
        }

        try:
            response = requests.post(smart_process_url, json=lead_data)
            response.raise_for_status()  # выбросить исключение для ответа с ошибкой
            
            logging.info("Смарт-процесс '%s' успешно создан!", item['TITLE'])
        except requests.exceptions.HTTPError as http_error:
            logging.error("Ошибка HTTP при создании смарт-процесса: %s, статус код: %d", http_error, http_error.response.status_code)
        except requests.exceptions.RequestException as req_error:
            logging.error("Ошибка запроса при создании смарт-процесса: %s", req_error)
        except Exception as e:
            logging.error("Неожиданная ошибка при создании смарт-процесса: %s", e)

      
def main_candidate_data():
    """ Основная функция которая получает информацию о id кандидата, производит выгрузку данных в
    функции get_candidate_data. 
    Резлуьтат записывается через функцию save_candidate_to_excel в файл формата .xlsx и ссылка 
    прикрепляется к карточки кандидата - save_link_to_file/
    """
    candidate_id = int(input("Напишите id кандидата: "))
    
    # получаем данные по кандидату
    candidate_data = get_candidate_data(candidate_id)
    if candidate_data:
        # генерируем Excel-файл
        excel_file = save_candidate_to_excel(candidate_data)
        print(f'Создание Excel-файла {excel_file} завершено!')
    
    # создания поля для прикрипления ссылки
    field_id = upload_file_to_lead('candidate_4_20241108_1700.xlsx')
    if field_id:
        save_link_to_file(field_id, 'candidate_4_20241108_1700.xlsx', candidate_id)
        print(f'Прикрепление карточки к кандидату завершено!')
        
 
def main_smart_process(): 
    """ Основная функция которая получает информацию файле формата .xlsx, который открывается для 
    чтения через функцию read_from_excel и из полученного результата создает смарт процесс
    с помощью функции create_smart_process 
    """
    data = read_from_excel('test_crm.xlsx')
    if data:
        create_smart_process(data)
        print('Добавление смарт процессов завершено!')


if __name__ == '__main__':
    main_candidate_data()
    main_smart_process()
